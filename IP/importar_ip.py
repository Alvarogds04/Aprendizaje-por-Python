import pyodbc
import openpyxl

cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=ZEATCSQL2.zena.local;DATABASE=VIPS;UID=CodisysSupport;PWD=Cambiame01')
cursor = cnxn.cursor()

query = "select PCDES from manhtv$$ where des ='TPV1'"
cursor.execute(query)

# Crear una nueva hoja de cálculo
wb = openpyxl.Workbook()
ws = wb.active

# Obtener los nombres de las columnas y escribirlos en la primera fila
columns = [column[0] for column in cursor.description]
for col_num, column_title in enumerate(columns, 1):
    ws.cell(row=1, column=col_num, value=column_title)

# Escribir los resultados de la consulta en las filas siguientes
for row_num, row in enumerate(cursor, 2):
    for col_num, cell_value in enumerate(row, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Guardar el archivo Excel
wb.save('resultados.xlsx')
